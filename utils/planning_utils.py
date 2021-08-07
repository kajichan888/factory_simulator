import sys
import random
import pandas as pd
import configparser
import pathlib
import numpy as np
import datetime
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

import jpholiday

from utils.utils import sec_to_timedelta_9h
from utils.utils import return_biz_day

path = pathlib.Path('__file__')
path /= '../'  # 1つ上の階層を指す
sys.path.append(str(path.resolve()))

config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

df = config['DEFAULT']['plan_dataframe']
df_path = pathlib.Path('__file__')
df_path /= '../'  # 1つ上の階層を指す
df_path = df_path.resolve() / df

def display_plan():
    """デフォルトファイルに登録されているの計画(DataFrame)を表示する"""
    df_from_pickle = pd.read_pickle(df_path)
    print(df_from_pickle)


def update_plan(products_dataframe):
    """DataFrameの計画をデフォルトにファイルに保存する"""
    products_dataframe.to_pickle(df_path)


def load_plan():
    """デフォルトファイルに登録されているの計画(DataFrame)を読み込む"""
    df_from_pickle = pd.read_pickle(df_path)
    return df_from_pickle


def replace_list(_list, index_number, m):
    """リストの順番を入れ替える関数
    
    _list: 順番を入れ替えるリスト
    index_number: 入れ替える場所のインデックス番号
    m: 動かす数 .. 前に1つなら -1 後ろに1つ動かすなら1
    
    >>> sample_index = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    >>> replace_list(sample_index, 4, -1)
    >>> sample_index
    [0, 1, 2, 4, 3, 5, 6, 7, 8, 9]
    
    >>> sample_index = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    >>> replace_list(sample_index, 4, -10)
    >>> sample_index
    [4, 0, 1, 2, 3, 5, 6, 7, 8, 9]
    
    >>> sample_index = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    >>> replace_list(sample_index, 4, 10)
    >>> sample_index
    [0, 1, 2, 3, 5, 6, 7, 8, 9, 4]
    
    """
    n = index_number + m
    if index_number + m < 0:
        n = 0
    elif index_number + m >= len(_list):
        n = len(_list)
    _list.insert(n, _list.pop(index_number))



def replace_plan(df, index_number, m):
    """DataFrameの順番を入れ替える関数
    
    df: 順番を入れ替えるDataFrame
    index_number: 入れ替える場所のインデックス番号
    m: 動かす数 .. 前に1つなら -1 後ろに1つ動かすなら1
    """
    lst = [i for i in range(len(df))]
    replace_list(lst, index_number, m)
    df = df.reindex(index=lst).reset_index(drop=True)
    
    return df

def shuffle_plan(df, index_number1, index_number2):
    """DataFrameの順番をシャッフルする関数
    
    df: 順番をシャッフルするDataFrame
    index_number1, 2 : シャッフルするレンジのインデックス番号 
    """
    lst = []
    l = [i for i in range(len(df))]
    for i in l[:index_number1]:
        lst.append(i)
    sample = random.sample(l[index_number1:index_number2+1], len(l[index_number1:index_number2+1]))
    for i in sample:
        lst.append(i)
    for i in l[index_number2+1:]:
        lst.append(i)
    df = df.reindex(index=lst).reset_index(drop=True)
    
    return df

def remove_plan(product_dataframe, df_row_label):
    """DataFrameから指定した列を削除する"""
    pr = product_dataframe.drop(df_row_label).reset_index(drop=True)
    return pr

def proceed_process_from_dataframe(product_dataframe, product_index_number):
    """
    pd.DataFrameで定義されたproductリストの中のproductの
    工程を一つ終わらせる

    procuct_dataframe: 編集するproductリスト(DataFrame)
    product_index_number: 工程を一つ終わらせるproductのインデックス番号
    """
    df = product_dataframe
    # 終わらせた工程をNaNで埋める
    for i in range(5, 9):
        df.iat[product_index_number, i] = np.nan
        
    lst = df.values.tolist()
    
    drop_list = []
    for l in lst:
        l_ = [i for i in l if str(i) != 'nan']
        # l_ = [i for i in l_ if i != None]
        drop_list.append(l_)
        
    max_processes = ''
    for _products in drop_list:
        if len(_products) > len(max_processes):
            max_processes = _products
    columns = ['product', 'lot', 'type', 'diameter', 'length']
    for i in range(1, int((len(max_processes) - 5) / 4 + 1)):
        columns.append(f'process-{i}')
        columns.append(f'time-{i}')
        columns.append(f'auto-{i}')
        columns.append(f'repeat-{i}')

    df = pd.DataFrame(drop_list, columns=tuple(columns))
    df = df.replace({None: float('nan'), 'nan': float('nan')})
    return df

def save_plan_to_excel(factory_obj, bar_color, BASE_DATE, file_path):
    """
    工程表をエクセルに変換して保存する
    
    factory_obj:シミュレート済みのFactory_obj
    BASE_DATE: 工程表の開始日
    file_path: 出力するエクセルファイルのパス
    """
    factory = factory_obj
    plan_file = file_path
    
    # Excelファイル新規作成
    wb = openpyxl.Workbook()

    # シート名を計画開始日にする
    sheet = wb.active
    sheet.title = str(BASE_DATE.date())

    # セルの最大値を決める
    max_col = (
        return_biz_day(BASE_DATE, sec_to_timedelta_9h(factory.time).days) - BASE_DATE
    ).days + 10

    plan = defaultdict(list)
    for product in factory.product_list:
        for fin in product.finished_process:
            plan[fin[0][0]].append([product.product,fin])

    canvas_0_cell = sheet['B5']

    r = canvas_0_cell.row

    for k, v in plan.items():
        r += 1
        for process in v:
            r += 2
    max_row = r

    # ===画面設定===

    # 列の幅=3.5
    for i in range(max_col):
        col = get_column_letter(i+1)
        sheet.column_dimensions[col].width=3.5

    # A列の幅=9
    # 幅をn文字に指定
    sheet.column_dimensions['A'].width = 9

    # 行の高さ=17.3
    for i in range(max_row):
        sheet.row_dimensions[i+1].height = 17.3
    # 1行目の高さ=31.5
    sheet.row_dimensions[1].height = 31.5

    # 2行目, 3行目 2列目以降->下罫線('細')　複数セルの取得(R1C1形式)
    cells = sheet.iter_rows(
        min_row = 2, max_row = 3,
        min_col = 2, max_col = max_col
    )
    for tpl in cells:
        for cell in tpl:
            cell.border = Border(
                bottom = Side(style='thin', color='000000',),
            )

    # A列右側に罫線(細)
    cells = sheet.iter_rows(
        min_row = 1, max_row = max_row,
        min_col = 1, max_col = 1
    )
    for tpl in cells:
        for cell in tpl:
            cell.border = Border(
                right = Side(style='thin', color='000000',),
            )

    # 4行目下罫線('細') 複数セルの取得(R1C1形式)
    cells = sheet.iter_rows(
        min_row = 4, max_row = 4,
        min_col = 1, max_col = max_col
    )
    for tpl in cells:
        for cell in tpl:
            cell.border = Border(
                bottom = Side(style='thin', color='000000',),
            )

    # 'A4' 右、下罫線
    sheet['A4'].border = Border(
        bottom = Side(style='thin', color='000000',),
        right= Side(style='thin', color='000000',),
    )

    # A1A2,A2A3結合
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('A3:A4')

    # 計画書タイトル
    cell1 = sheet['A1']
    cell_str1 = '計画書No.' + '\n' + str(BASE_DATE.date()).replace('-', '')
    cell1.value = cell_str1

    cell2 = sheet['A3']
    cell_str2 = BASE_DATE.strftime('%Y/%m') + '\n' + '->'
    cell2.value = cell_str2

    cells = [cell1, cell2]
    # フォントの設定
    for cell in cells:
        cell.font = openpyxl.styles.Font(
            name = "ＭＳ Ｐゴシック",
            size = 10,
        )
        cell.alignment = Alignment(
            horizontal = 'center',
            vertical = 'center',
            wrapText = True,
        )

    # === 日付、曜日の挿入 ===

    start_cell = sheet['B3']

    week = ['月', '火', '水', '木', '金', '土', '日']
    # WeekNo.2の月曜日
    w2 = datetime.datetime(2021, 1, 4)
    # 土曜出勤設定
    working_saturday = [datetime.datetime(2021,2,13), datetime.datetime(2021,2,14)]

    d = BASE_DATE
    r = start_cell.row
    c = start_cell.column
    for i in range(max_col):
        # 曜日
        sheet.cell(row=r+1, column=c).value = week[d.weekday()]
        # 日付
        if i == 0 or d.day ==1 :
            sheet.cell(row=r, column=c).value = d.strftime('%m/')
        else:
            sheet.cell(row=r, column=c).value = d.strftime('%d')
        # 月曜日weekday表示
        if d.weekday() == 0:
            sheet.cell(row=r-1, column=c).value = 'W' + str(int((d-w2).days/7 + 2))
            sheet.cell(row=r-1, column=c).font = openpyxl.styles.Font(size = 10)
        # 土日祝日塗りつぶし
        if (d.weekday() >=5 or jpholiday.is_holiday(d)) and d not in working_saturday:
            sheet.cell(row=r+1, column=c).font = openpyxl.styles.Font(color='FF0000')
            for n in range(max_row):
                cell = sheet.cell(row=n+1, column=c)
                cell.fill = openpyxl.styles.PatternFill(
                    patternType = 'solid',
                    fgColor = 'D3D3D3',
                )
        d += datetime.timedelta(days=1)
        c += 1

    # === 工程表作成 ===

    plan = defaultdict(list)
    for product in factory.product_list:
        for fin in product.finished_process:
            plan[fin[0][0]].append([product.product,fin])

    canvas_0_cell = sheet['B5']

    r = canvas_0_cell.row
    c = canvas_0_cell.column

    for k, v in plan.items():
        # 列の高さを低く
        sheet.row_dimensions[r].height = 5

        v = sorted(v, key=lambda x:x[1][1])
        r += 1
        sheet.cell(row=r, column=c-1).value = k

        for process in v:
            date_range = (
                return_biz_day(BASE_DATE, sec_to_timedelta_9h(process[1][2]).days)
                - return_biz_day(BASE_DATE, sec_to_timedelta_9h(process[1][1]).days)
            ).days
            for n in range(date_range + 1):
                start_date =(
                    return_biz_day(BASE_DATE, sec_to_timedelta_9h(process[1][1]).days)
                    - BASE_DATE
                )
                cell = sheet.cell(
                    row = r, 
                    column = c + start_date.days + n,
                )
                cell.fill = openpyxl.styles.PatternFill(
                    patternType = 'solid',
                    fgColor = bar_color,)

            product_name_cell = sheet.cell(
                row = r,
                column= c + start_date.days + n + 1,
            )
            product_name_cell.value = process[0]
            r += 1

            # 列の高さを低く
            sheet.row_dimensions[r].height = 5
            r += 1

        # 罫線
        cells = sheet.iter_rows(
        min_row = r, max_row = r,
        min_col = 1, max_col = max_col
        )
        for tpl in cells:
            for cell in tpl:
                cell.border = Border(
                    top = Side(style='thin', color='000000',),
                )
        # '機械名欄 右、上罫線
        sheet.cell(row=r, column=1).border = Border(
        top = Side(style='thin', color='000000',),
        right= Side(style='thin', color='000000',),
        )
        # 列の高さを低く
        sheet.row_dimensions[r].height = 5
    
    # 保存
    wb.save(plan_file)
    wb.close()

if __name__ == '__main__':
    import doctest
    doctest.testmod()
